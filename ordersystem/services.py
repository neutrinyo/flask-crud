from enum import Enum
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


class ColorEnum(Enum):
    '''
    ColorEnum - inherits from the Enum class. Enumerates status names into relevant PatternFill objects.
    Allows for the comfortable conversion of status string into cell color - that is why it is not conforming
    to the uppercase style standard.

    Fields:
    - New - PatternFill object representing a light blue solid color.
    - Inprogress - PatternFill object representing a yellow solid color.
    - Completed - PatternFill object representing a light green solid color.
    '''
    New = PatternFill(start_color='82C9FF', end_color='82C9FF', fill_type='solid')
    Inprogress = PatternFill(start_color='FDFF73', end_color='FDFF73', fill_type='solid')
    Completed = PatternFill(start_color='65EB91', end_color='65EB91', fill_type='solid')



def parse_bulk_order_inputs(request_form: dict) -> list:
    '''
    parse_bulk_order_inputs - business logic function. Takes a request form dictionary with id and status lists in string form, 
    parses them into two separate lists, converts the id numbers to int. 
    Checks whether the lists are equal in length. Returns the parsed lists.

    Input 
    - request_form - dictionary with two required fields: id_list and status_list. id_list has to be a string of ints separated by commas
    (without whitespace), status_list has to be a comma separated (without whitespaces after commas) list of valid status strings as described
    in models.py's Order class. (New, In progress, Completed)

    Output:
    - id_numbers, status_list - lists of validated id ints and status strings. In case of an error, they are an error message and an HTTP status code.
    '''
    try:
        id_list = request_form['id_list'].split(',')
        status_list = request_form['status_list'].split(',')
        id_numbers = [int(order_id) for order_id in id_list]
    except KeyError as ke:
        return f"Missing required argument: {ke}", 400
    except ValueError as ve:
        return f"One of the input lists has an invalid value: {ve}", 400

    if len(id_list) != len(status_list):
        return "Unequal lengths of status and id lists.", 400
    
    return id_numbers, status_list



def compute_statistics(order_dict: list) -> str:
    '''
    get_statistics - function called for the /orders/stats URL.
    Takes a list of dictionaries of all orders, loads it into a Pandas DataFrame and computes the following statistics: 
    total entry number, oldest entry, newest entry, number of entries per status.

    Input: none.
    Output: format string with the statistics mentioned above.
    '''

    order_df = pd.DataFrame.from_dict(order_dict)
    
    status_count = order_df.groupby(['status']).size().reset_index(name="count").to_string()
    database_size = len(order_df)

    oldest_entry = order_df[order_df['creation_date'] == order_df['creation_date'].min()]
    oldest_entry_id = oldest_entry['id'].values[0]
    oldest_entry_date = oldest_entry['creation_date'].values[0]
    oldest_entry_date = pd.to_datetime(str(oldest_entry_date))
    oldest_entry_date = oldest_entry_date.strftime('%Y.%m.%d, %H:%M')

    newest_entry = order_df[order_df['creation_date'] == order_df['creation_date'].max()]
    newest_entry_id = newest_entry['id'].values[0]
    newest_entry_date = newest_entry['creation_date'].values[0]
    newest_entry_date = pd.to_datetime(str(newest_entry_date))
    newest_entry_date = newest_entry_date.strftime('%Y.%m.%d, %H:%M')

    return database_size, oldest_entry_id, oldest_entry_date, newest_entry_id, newest_entry_date, status_count



def generate_xlsx_report(order_dict: list) -> Workbook:
    '''
    generate_xlsx_report - function called for the /orders/report URL.
    Takes a list of dictionaries of all orders, then turns it into a DataFrame which is then slotted into an empty Workbook object. 
    Each row is then colored in accordance to its status, as described in ColorEnum. 

    Input: order_dict - list of dicts returned by list_all_orders.
    Output: A Workbook object with all previously mentioned features.
    '''
    workbook = Workbook()
    worksheet = workbook.active
    order_df = pd.DataFrame.from_dict(order_dict)

    # conversion of DateTime objects into strings.
    order_df['creation_date'] = order_df['creation_date'].astype('str')
    rows = dataframe_to_rows(order_df, index=False, header=True)

    for row in rows:
        worksheet.append(row)

    for row in worksheet.iter_rows(min_row=2):
        status = row[2].value.replace(" ", "")
        style = ColorEnum[status].value

        for cell in row:
            cell.fill = style

    return workbook